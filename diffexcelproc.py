import os
import openpyxl as xld
import csv
import diffexcelconfig as config

class DiffExcelProc :
    #コンストラクタ
    def __init__(self, first ,second):
        self.first_filepath = first
        self.second_filepath = second

    #Excelファイル読み込み    
    def get_file_data(self):
        
        #ファイル存在する場合、ファイルを読み込む
        if os.path.isfile( self.first_filepath) and os.path.isfile( self.second_filepath):
            self.first_workbook = xld.load_workbook( self.first_filepath,data_only = True)
            self.second_workbook = xld.load_workbook(self.second_filepath, data_only = True)
            #シートリスト取得
            self.first_sheetnames = self.first_workbook.sheetnames
            self.second_sheetnames = self.second_workbook.sheetnames
            #ファイルパスからファイル名のみ抽出
            self.first_filename = os.path.basename(self.first_filepath) 
            self.second_filename = os.path.basename(self.second_filepath)
        else:
            return False
        
        return True

    #シート名取得
    def get_sheet_name(self):

        # 同一シート名のみ抽出
        self.sheet_names = []
        for first_name in self.first_sheetnames:
            for second_name in self.second_sheetnames:
               if first_name == second_name:
                self.sheet_names.append(first_name)


        #同一シート名のシートが存在しない場合
        if len(self.sheet_names) == 0:
            return False

        return True
        

    #差分比較
    def diff_sheet( self, first_sheetname, second_sheetname, selected_mode):
        first = self.first_workbook[first_sheetname]
        second = self.second_workbook[second_sheetname]

        diff_count = 0
        
        #行・列のサイズが大きい方を取得
        sheet_max_clm = max(first.max_column, second.max_column)
        sheet_max_row = max(first.max_row, second.max_row)
        
        for  j in range(1 , sheet_max_row + 1):
            for i in range( 1 , sheet_max_clm + 1):
                if first.cell(row = j, column = i).coordinate is None:
                    cell_addr = second.cell(row = j, column = i).coordinate
                else:
                    cell_addr = first.cell(row = j, column = i).coordinate
                
                #空白・空文字・スペースセルを除去
                first_value = self.padding_space(first[cell_addr].value)
                second_value = self.padding_space(second[cell_addr].value)
                #差分ありの場合、リストに格納
                if first_value != second_value: 
                    #output用リストに結合
                    if selected_mode == config.ALLSHEET_MODE:
                        diff_col = [first_sheetname, cell_addr, first_value, second_value]
                    elif selected_mode == config.SELECTSHEET_MODE:
                        diff_col = [cell_addr, first_value, second_value]
                    self.diff_list.append(diff_col)
                    diff_count = diff_count + 1
        return diff_count
                                      
    #空白・空文字・スペースセルを除去
    def padding_space(self, target):
        if target is None:
            #return None
            return ''
        #空文字
        elif len(str(target)) == 0:
            #return None
            return ''
        #スペースのみ
        elif str(target).isspace == True:
            return ''
        else:
            return target
    
    #差分抽出
    def output_diff(self, selected_mode, first_sheet = '', second_sheet = ''):
        rtn = 0
        
        if selected_mode == config.ALLSHEET_MODE:
            #ヘッダー作成
            self.diff_list = [[config.SHEETNAME_COL,  config.CELLADDR_COL, self.first_filename, self.second_filename]]
            #シート毎に差分抽出
            for sheets in range(len(self.sheet_names)):
                rtn = rtn + self.diff_sheet(self.sheet_names[sheets], self.sheet_names[sheets] , config.ALLSHEET_MODE)
        elif selected_mode == config.SELECTSHEET_MODE:
            #ヘッダー作成
            first_sheet_col = config.FILENAME_COL + '：' + self.first_filename + '\n' + config.SHEETNAME_COL + '：' + first_sheet
            second_sheet_col = config.FILENAME_COL + '：' + self.second_filename + '\n' + config.SHEETNAME_COL + '：'+ second_sheet
            self.diff_list = [[config.CELLADDR_COL, first_sheet_col, second_sheet_col]]
            #差分抽出
            rtn = rtn + self.diff_sheet( first_sheet, second_sheet, config.SELECTSHEET_MODE)
        else:
            return rtn

        #CSVファイルに出力
        with open('diff.csv', 'wt', encoding='utf-8-sig', newline='') as fout:
            # ライター（書き込み者）を作成
            writer = csv.writer(fout)
            # ライターでデータ（行列）をファイルに出力
            writer.writerows(self.diff_list)
        return rtn
